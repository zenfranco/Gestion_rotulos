import sqlite3,sys
import clipboard as c
from openpyxl import Workbook
from openpyxl.styles import Font
from querys import *
from PyQt4 import QtCore, QtGui, uic
from datetime import date
global rango, numpedido, pedidos, disponible,subpedidos,INICIAL,FINAL
#Rango =[1,1000] #[valor inicial, valor final]
pedidos=[]
subpedidos=[]



# Cargar archivo .ui
form_class = uic.loadUiType("main.ui")[0]

class VentanaPrincipal(QtGui.QMainWindow, form_class):
	def __init__ (self,parent=None):
		QtGui.QMainWindow.__init__(self, parent)
		self.setupUi(self)
		
		self.frame_detallepedido.hide()
		self.frm_manual.hide()
		#CAMBIAR DE PAGINAS
		self.btn_nuevopedido.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_nuevopedido)) #cambia de pagina
		self.btn_nuevosubpedido.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_nuevosubpedido)) #cambia de pagina
		self.btn_listar.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_listar)) #cambia de pagina
		self.home.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_inicio)) #cambia de pagina
		self.btn_rendicion.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_rendicion)) #cambia de pagina
		self.btn_configuracion.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_config)) #cambia de pagina
		self.btn_deposito.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_deposito)) #cambia de pagina
		self.btn_rotulos.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_rotulos)) #cambia de pagina
		self.btn_nuevagestion.clicked.connect(lambda: self.stackedWidget.setCurrentWidget(self.p_gestion)) #cambia de pagina
		
		#FUNCION DE LOS BOTONES
		#PAGINA PEDIDOS
		self.btn_iniciarpedido.clicked.connect(self.iniciarpedido)
		self.btn_ingresarpedido.clicked.connect(self.NuevoPedido)
		self.btn_ingresarpedido.clicked.connect(self.limpiar)
		self.combo_asociados.activated.connect(self.traeregistro)
		self.cbx_manual.stateChanged.connect(lambda:self.frm_manual.show())
		
		
			
		#boton salir
		self.btn_salir.clicked.connect(salir)
		
		#pagina subpedidos
		self.btn_verpedidos.clicked.connect(self.verpedidos)
		self.btn_subvalidar.clicked.connect(self.validarpedido)
		self.btn_subingresar.clicked.connect(self.nuevosubpedido)
		self.btn_subingresar.clicked.connect(self.limpiar)
		self.combo_asociados_subpedidos.activated.connect(self.traeregistrosp)
		self.tb_verpedidos.itemDoubleClicked.connect(self.completanumpedido)
		self.btn_copy_inicio.clicked.connect(self.clipinicio)
		self.btn_refresh.clicked.connect(self.refresh_pedidos)
		
		
		#pagina listar
		self.btn_listarlistar.clicked.connect(self.listar)
		self.combo_asociados_listar.activated.connect(self.traeregistrolistar)
		self.btn_pedidoexcell.clicked.connect(self.listartoexcell)
		
		#PAGINA RENDICION
		self.btn_consultar.clicked.connect(self.rendir)
		self.btn_rendir.clicked.connect(self.exportarrendicion)
		
		#PAGINA CONFIGURACION
		self.btn_definir_locker.clicked.connect(self.setearlockers)
		self.btn_cargarasociado.clicked.connect(self.altasocio)
		self.btn_agregarrango.clicked.connect(self.agregarrango)
		self.btn_ver_rangos.clicked.connect(self.verrangos)
		self.btn_definir_rango.clicked.connect(self.setearrango)
		
		#PROPIEDADES
		self.txt_indice_rotulos.hide()
		self.rb_porrncyfs.setChecked(True)
		self.rb_todos.setChecked(True)
		self.fechadesde_rendicion.setDate(date.today())
		self.fechahasta_rendicion.setDate(date.today())
		self.fecha_desde_rotulos.setDate(date.today())
		self.fecha_hasta_rotulos.setDate(date.today())
		self.fecha_desde_listar.setDate(date.today())
		self.fecha_hasta_listar.setDate(date.today())
		self.signal_gestion_indice.hide()
		self.cbx_porfecha.stateChanged.connect(lambda:self.fecha_desde_listar.setEnabled(True))
		self.cbx_porfecha.stateChanged.connect(lambda:self.fecha_hasta_listar.setEnabled(True))
		
		#AJUSTAR CONTENIDO DE LAS CELDAS
		headertb_gestiones = self.tb_gestiones.horizontalHeader()
		headertb_gestiones.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		
		headertb_verpedidos = self.tb_verpedidos.horizontalHeader()
		headertb_gestiones.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		
		headertb_lockers = self.tb_lockers.horizontalHeader()
		headertb_lockers.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		
		headertb_rangos = self.tb_rangos.horizontalHeader()
		headertb_rangos.setResizeMode(QtGui.QHeaderView.ResizeToContents)
		
		
		
		
		#PAGINA DEPOSITO
		self.btn_almacenar.clicked.connect(self.almacenar)
		self.tb_lockers.itemDoubleClicked.connect(self.lockerselected)
		self.btn_despachar.clicked.connect(self.despachar)
		self.btn_deposito.clicked.connect(self.listarlockers)
		self.btn_buscar_locker.clicked.connect(self.filtrarlockers)
		
		#PAGINA ROTULOS
		self.btn_rotulos.clicked.connect(self.listarimpresiones)
		self.btn_ingresar_rotulos.clicked.connect(self.nuevaimpresion)
		self.cb_razonsocial_rotulos.activated.connect(self.traeregistrorotulos)
		self.btn_listarimpresiones.clicked.connect(self.listarimpresiones)
		self.btn_definir_rotulos.clicked.connect(self.cambiarestadorotulo)
		self.tb_rotulos.itemDoubleClicked.connect(self.impresionselected)
		
		#PAGINA GESTIONES
		self.btn_agregar_nueva.clicked.connect(self.nuevagestion)
		self.btn_buscar_gestiones.clicked.connect(self.listargestiones)
		self.tb_gestiones.itemDoubleClicked.connect(self.gestionselected)
		self.btn_actualizargestion.clicked.connect(self.gestionupdate)
		self.btn_nuevagestion.clicked.connect(self.listargestiones)
		self.btn_buscar_gestiones_filtro.clicked.connect(self.filtrargestiones)
		self.btn_ingresarnota.clicked.connect(self.notaupdate)
		
		

	
		
	def llenarcombo(self):
		q=bdquery()
		asociados = q.traerasociados() #la consulta devuelve una tupla, por lo tanto hay que convertirla a str para llenar el combobox, se usa el metodo "".join()
		
		k=0
		for i in asociados:
			self.combo_asociados.addItem("".join(asociados[k]))
			self.combo_asociados_subpedidos.addItem("".join(asociados[k]))
			self.combo_asociados_listar.addItem("".join(asociados[k]))
			self.cb_razonsocial_rotulos.addItem("".join(asociados[k]))
			self.combo_asociados_gestiones.addItem("".join(asociados[k]))
			k=k+1
		
		
		
	def iniciarpedido(self):
		global INICIAL
		global FINAL
		global Numpedido
		global indice
		q = bdquery()
		dato= q.traeultimopedido()
		
		if self.rb_seriea.isChecked():
			indice=1
		else:
			indice=2
				
		rangogeneral = q.recuperarango(indice)
		
		Numpedido=dato[0]
		INICIAL=int(rangogeneral[0])
		FINAL=int(rangogeneral[1])
		
		
		disponible = int(FINAL-INICIAL+1)
		self.signal_stock.setText(str(disponible))
		self.signal_inicial.setText(str(INICIAL))
		self.signal_final.setText(str(FINAL))
		
				
		
	def traeregistro(self):
		nombre = str(self.combo_asociados.currentText())
		
		q=bdquery()
		recuperado=q.getrncyfs(nombre)
		
			
		self.txt_rncyfs.setText("".join(recuperado))
	def traeregistrosp(self):
		nombre = str(self.combo_asociados_subpedidos.currentText())
		
		q=bdquery()
		recuperado=q.getrncyfs(nombre)
		
			
		self.txt_verpedido.setText("".join(recuperado))
	def traeregistrolistar(self):
		nombre = str(self.combo_asociados_listar.currentText())
		
		q=bdquery()
		recuperado=q.getrncyfs(nombre)
		
			
		self.txt_listar.setText("".join(recuperado))
	
	def traeregistrorotulos(self):
		nombre = str(self.cb_razonsocial_rotulos.currentText())
		
		q=bdquery()
		recuperado=q.getrncyfs(nombre)
		
			
		self.txt_rncyfs_rotulos.setText("".join(recuperado))
	
	def traeregistrogestiones(self):
		nombre = str(self.combo_asociados_gestiones.currentText())
		
		q=bdquery()
		recuperado=q.getrncyfs(nombre)
		
			
		return ("".join(recuperado))
		
		
	def NuevoPedido(self):
		
		if self.txt_cantidad.text():
			
		
			global INICIAL
			global FINAL
			global Numpedido
			global indice
					
		
			disponible = FINAL-INICIAL+1
		
			cantidad = int(self.txt_cantidad.text())
			registro = str(self.txt_rncyfs.text())
		
			if cantidad <= disponible:

				P = Pedido(cantidad,registro,Numpedido) #crea nuevo pedido
				P.asignar(INICIAL) #envia como parametro el rango inicial

			
			
				P.showrango()
				pedidos.append(P)
				Numpedido= Numpedido+1
				q= bdquery()
				estado="SIN USAR"
				fechapedido=str(date.today())
				if self.rb_seriea.isChecked():
					serie="A"
				else:
					serie="B"
				
				
				
				q.cargapedido(Numpedido,registro,cantidad,INICIAL,INICIAL+cantidad-1,INICIAL,INICIAL+cantidad-1,estado,fechapedido,serie)
				
				
				q.incrementanpedido(Numpedido)
				self.signal_pedidoexitoso.setText("Pedido Creado")
				
				self.frame_detallepedido.show()			
				self.signal_cantidad.setText(str(cantidad))
				self.signal_rncyfs.setText(str(registro))
				self.signal_inicio.setText(str(INICIAL))
				self.signal_fin.setText(str(INICIAL+cantidad-1))
				self.signal_numpedido.setText(str(Numpedido))
				
				#GUARDA DATOS PARA IMRIMIR
				nombre=q.traenombre(registro)
				
				ticket= open("ticket.txt","w")
				
				ticket.write("DETALLE DE PEDIDO\n")
				ticket.write("-----------------------------\n")
				ticket.write("RAZON SOCIAL: "+str("".join(nombre))+"\n")
				ticket.write("PEDIDO: "+str(Numpedido)+"\n")
				ticket.write("-----------------------------\n")
				ticket.write("RNCyFS: "+str(registro)+"\n")
				ticket.write("Fecha: "+str(fechapedido)+"\n")			
				ticket.write("Rango: ")
				ticket.write(str(INICIAL)+" - "+str(INICIAL+cantidad-1)+"\n")
				ticket.write("Serie: "+str(serie)+"\n")
				ticket.write("Cantidad: "+str(cantidad)+"\n")
				
				
				ticket.write("-----------------------------\n")
				ticket.close()
							
				
				
				
				INICIAL=INICIAL+cantidad
				q.actualizarangoenbd(INICIAL,FINAL,indice)
				self.combo_asociados.clear()
				self.llenarcombo()
								
			
			else:
				
				self.signal_pedidoexitoso.setText("No hay suficiente stock")
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("INGRESE CANTIDAD")
			msgBox.exec_()
			
			
	def limpiar(self):
		self.txt_cantidad.setText("")
		self.txt_rncyfs.setText("")
		self.txt_kg.setText("")
		
		self.txt_subcantidad.setText("")
		self.txt_subvariedad.setText("")
		
		
		
	
	def verpedidos(self):
		
		campo= str(self.txt_verpedido.text())
		q=bdquery()
		tablapedidos=q.verpedido(campo)
		totalfilas=len(tablapedidos)
		self.tb_verpedidos.setRowCount(totalfilas)
		
		fila=0
		
		
		for i in tablapedidos:
			
						
			self.tb_verpedidos.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_verpedidos.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_verpedidos.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_verpedidos.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))  
			
			
			fila=fila+1
		
	def completanumpedido(self):
		
		fila = self.tb_verpedidos.currentRow()
		pedido=self.tb_verpedidos.item(fila, 0).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		self.txt_numpedido.setText(pedido)
		self.validarpedido()
		
		
	def refresh_pedidos(self):
		self.validarpedido()
		self.verpedidos()	
				
	def validarpedido(self):
		numeropedido=int(self.txt_numpedido.text())
		q=bdquery()
		tablapedidos=q.getpedido(numeropedido)
		
		if len(tablapedidos) != 0:
			disponible=int(tablapedidos[7]-tablapedidos[6]+1)
		
			
			self.signal_disponibleinicio_sp.setText(str(tablapedidos[6]))
			self.signal_disponiblefin_sp.setText(str(tablapedidos[7]))
			
			self.signal_stock_sp.setText(str(disponible))
		
		
			
	def clipinicio(self):
		inicio=int(self.signal_disponibleinicio_sp.text())
		c.copy(inicio)
		
		
	def nuevosubpedido(self):
		
		
		if self.txt_subvariedad.text() and self.txt_subcantidad.text() and self.txt_kg.text() and self.cbx_especie_sp.currentText() and self.cbx_categoria_sp.currentText(): 
		
			numeropedido=int(self.txt_numpedido.text())
			q=bdquery()
			tablapedidos=q.getpedido(numeropedido)
			cantidad =int(self.txt_subcantidad.text())
			
			if len(tablapedidos) != 0:
				disp=int(tablapedidos[7]-tablapedidos[6]+1)
				
				if disp >= cantidad:
					
					registro=str(tablapedidos[1])
					
					if self.cbx_manual.isChecked(): #OPCION PARA QUE EL USUARIO INGRESE SU RANGO DE INICIO
						
						rangomanual=self.txt_manual.text()
						
						if rangomanual in range(tablapedidos[6],tablapedidos[7]):
							
							spini=rangomanual
							spfin=spini+cantidad-1
							
								
							
						else:
							print "rango incorrecto"
							
						
						#VALIDAR QUE NO SUPERE EL RANGO FINAL
						
						
						
						
					
					
					else:#CAMINO AUTOMATICO: EL SISTEMA ASIGNA SECUENCIALMENTE EL RANGO
						
						spini=int(tablapedidos[6]) #valor inicioremanente de pedido
						spfin=spini+cantidad-1
					
					numpedido=int(self.txt_numpedido.text())
					variedad=str(self.txt_subvariedad.text())
					especie=str(self.cbx_especie_sp.currentText())
					KG=int(self.txt_kg.text())
					categoria=str(self.cbx_categoria_sp.currentText())
					camp=self.txt_subcamp.text()
					fechasubpedido=str(date.today())
					
					
					
					if variedad=="":
						variedad="n/d"
					if especie=="n/d":
						especie="n/d"
					
					if categoria=="":
						categoria="n/d"
					if KG=="":
						dav=0
						
					ticket= open("subpedido.txt","a")
				
					
					ticket.write(str(cantidad)+" ")		
					ticket.write(str(variedad)+" ")
					ticket.write(str(spini)+" - "+str(spini+cantidad-1)+"\n")
					
				
					
					ticket.close()
					
					
				
					'''#EXPORTA A ARCHIVO EXCELL
					book = Workbook()
					sheet = book.active
					
					
					sheet['I13']=str(spini)+" - "+str(spini+cantidad-1)
				
					
				
				
					book.save('nota_davs.xls',"a")'''
					
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(1)
					msgBox.setWindowTitle("SUBPEDIDO")
					msgBox.setText("SUBPEDIDO CREADO")
					msgBox.exec_()
					
					c.copy(str(spini)+"-"+str(spfin))
					q=bdquery()
					q.cargasubepedido(numpedido,spini,spfin,cantidad,variedad,especie,int(camp),int(KG),categoria,registro,fechasubpedido)
					
					inicioremanente=tablapedidos[6]+cantidad
					disponible=tablapedidos[7]-inicioremanente+1
					if disponible <=0:
						estado="FINALIZADO"
					else:
						estado="VIGENTE"
						
					q.actualizaremanente(numpedido,inicioremanente)				 #se actualiza el stock remanente del pedido en tabla pedidos
					q.actualizaestado(numpedido,estado)
					self.refresh_pedidos()
				else:
					msgBox=QtGui.QMessageBox(self.centralwidget)
					msgBox.setIcon(3)
					msgBox.setWindowTitle("ERROR DE STOCK")
					msgBox.setText("NO HAY STOCK SUFICIENTE PARA ESTA SOLICITUD")
					msgBox.exec_()
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("HAY CAMPOS VACIOS")
			msgBox.exec_()
				
				
				
			
	def altasocio(self):
		registro =str(self.txt_altaasociado_reg.text())
		nombre =str(self.txt_altaasociado_nombre.text())
		q=bdquery()
		q.altaasociado(registro,nombre)
		self.txt_altaasociado_reg.setText("")
		self.txt_altaasociado_nombre.setText("")
		
		
		
	def nuevagestion(self):
		
		if self.txt_cantidad_gestiones.text():
		
			registro =self.traeregistrogestiones()
			tipo =str(self.combo_gestiones.currentText())
			cantidad =int(self.txt_cantidad_gestiones.text())
			estado="0-Iniciado"
			
			q=bdquery()
			q.altagestion(registro,tipo,estado,date.today(),cantidad)
			
			self.combo_gestiones.setCurrentIndex(0)
			self.combo_asociados_gestiones.setCurrentIndex(0)
			self.txt_cantidad_gestiones.setText("")
			self.listargestiones()
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("INGRESE CANTIDAD")
			msgBox.exec_()
		
		
		
		
	def listargestiones(self):
		
		nombre= "%"
		if self.rb_gestiones_pendientes.isChecked():
			estado="%-%"
		elif self.rb_gestiones_finalizadas.isChecked():			
			estado="Finalizado"
		elif self.rb_gestiones_todas.isChecked():	
			estado="%"
		
		
		q=bdquery()
		
		listarecuperada=q.traergestiones(estado,nombre)
		totalfilas=len(listarecuperada)
		self.tb_gestiones.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_gestiones.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_gestiones.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_gestiones.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_gestiones.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_gestiones.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_gestiones.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			self.tb_gestiones.setItem(fila,6,QtGui.QTableWidgetItem(str(i[6])))
									
			fila=fila+1
		self.signal_gestiones.setText(str(totalfilas))
		
	def filtrargestiones(self):
		if str(self.combo_estados_gestiones_filtro.currentText()) =="-":
			estado="%"
		else:
			estado=str(self.combo_estados_gestiones_filtro.currentText())
			
		asociado=str("%"+self.txt_asociados_gestiones_filtro.text()+"%")
	
		q=bdquery()
		
		listarecuperada=q.traergestiones(estado,asociado)
		totalfilas=len(listarecuperada)
		self.tb_gestiones.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_gestiones.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_gestiones.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_gestiones.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_gestiones.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_gestiones.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_gestiones.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			self.tb_gestiones.setItem(fila,6,QtGui.QTableWidgetItem(str(i[6])))
									
			fila=fila+1
		self.signal_gestiones.setText(str(totalfilas))
		
			
	def gestionselected(self):
		fila = self.tb_gestiones.currentRow()
		indice=self.tb_gestiones.item(fila, 6).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		registro=self.tb_gestiones.item(fila, 0).text()
		fecha=self.tb_gestiones.item(fila, 5).text()
		estado=self.tb_gestiones.item(fila, 1).text()
		
		q=bdquery()
		obs=q.traenotas(int(indice))
		
		self.signal_gestion_asociado.setText(str(registro))
		self.signal_gestion_fecha.setText(str(fecha))
		self.signal_gestion_estado.setText(str(estado))
		self.signal_gestion_indice.setText(str(indice))
		self.signal_gestion_observaciones.setText(str("".join(obs)))
		
	def gestionupdate(self):
		indice=int(self.signal_gestion_indice.text())
		estado=str(self.combo_estados_gestiones.currentText())
		q=bdquery()
		q.updategestion(indice,estado)
		
		self.combo_estados_gestiones.setCurrentIndex(0)
		self.listargestiones()
		
	def notaupdate(self):
		indice=int(self.signal_gestion_indice.text())
		obs=str(self.signal_gestion_observaciones.toPlainText())
		q=bdquery()
		q.insertarnota(indice,obs)
		self.listargestiones()
		
			
	def listar(self):
		q=bdquery()
		
		if self.cbx_porrotulo.isChecked():
			rotulo=int(self.txt_porrotulo.text())
			tablarecuperada = q.busquedaxrotulo(rotulo)
			
			
		else:
			
		
				
			if self.rb_vigentes.isChecked():
				radiobutton="VIGENTE"
			elif self.rb_finalizados.isChecked():
				radiobutton="FINALIZADO"
			elif self.rb_todos.isChecked():
				radiobutton="%"			
			elif self.rb_sinusar.isChecked():
				radiobutton="SIN USAR"
				
			desde=str(self.fecha_desde_listar.text())
			hasta=str(self.fecha_hasta_listar.text())
		
		
		
			
			if self.rb_porrncyfs.isChecked():
				campolistar= str(self.txt_listar.text())
				
				if self.cbx_porfecha.isChecked():
					
					tablarecuperada=q.listaxregistrofecha(campolistar,radiobutton,desde,hasta)
				else:
					tablarecuperada=q.listaxregistro(campolistar,radiobutton)
				
				
				
			
				
			elif self.rb_pornumpedido.isChecked():
				campolistar= int(self.txt_listar.text())
				q=bdquery()
				if self.cbx_porfecha.isChecked():
					tablarecuperada=q.listaxpedidofecha(campolistar,radiobutton,desde,hasta)
				else:
					tablarecuperada=q.listaxpedido(campolistar,radiobutton)
				
			
		
		totalfilas=len(tablarecuperada)
		self.tb_listar.setRowCount(totalfilas)		
			
		fila=0
		
		acum=0
		for i in tablarecuperada:
			
						
			self.tb_listar.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_listar.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_listar.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2]))) #RAZON SOCIAL
			self.tb_listar.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_listar.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_listar.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			self.tb_listar.setItem(fila,6,QtGui.QTableWidgetItem(str(i[6])))
			self.tb_listar.setItem(fila,7,QtGui.QTableWidgetItem(str(i[7])))
			self.tb_listar.setItem(fila,8,QtGui.QTableWidgetItem(str(i[8])))
			self.tb_listar.setItem(fila,9,QtGui.QTableWidgetItem(str(i[9])))
			self.tb_listar.setItem(fila,10,QtGui.QTableWidgetItem(str(i[10])))
			self.tb_listar.setItem(fila,11,QtGui.QTableWidgetItem(str(i[11])))
			
		
			acum=acum+int(i[4])
				
			fila+=1
			
		self.signal_total_listar.setText(str(acum))
			
	def listartoexcell(self):
		
		book = Workbook()
		sheet = book.active
		
		
		q=bdquery()
		
		
		if self.rb_pornumpedido.isChecked():
			campolistar= int(self.txt_listar.text())
			tablarecuperada=q.listaraexcell(campolistar)
			#TRAER RAZON SOCIAL CON NUM DE PEDIDO O CON NUM DE REGISTRO
		else:
			campolistar= str(self.txt_listar.text())
			tablarecuperada=q.listaraexcellall(campolistar)
		
		
		sheet['A1']="ASOCIADO"
		sheet['B1']=""
		sheet['A2']="PEDIDO"
		sheet['B2']="CANTIDAD"
		sheet['C2']="RANGO UTILIZADO"
		sheet['D2']="VARIEDAD"
		sheet['E2']="ESPECIE"
		sheet['F2']="CATEGORIA"
		sheet['G2']="ENVASES/KG"
		
		sheet['H2']="CONTROL"
		
		for i in tablarecuperada:
			sheet.append(i)
			
					
						
			
		book.save('subpedido.xls')
			
		
	def rendir(self):
		if self.rb_todo_rendicion.isChecked():
			
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			especie=str(self.cbx_especie_rendicion.currentText())
			cultivar=str(self.cbx_cultivar_rendicion.currentText())
			categoria=str(self.cbx_categoria_rendicion.currentText())
			camp= str(self.cbx_camp_rendicion.currentText())
		
		
		
		
			if registro =="":
				registro="%"
			
			if especie =="":
				especie="%"
				
			if cultivar =="":
				cultivar="%"
		
			if categoria =="":
				categoria="%"
			
			if camp =="":
				camp="%"
			
			
				
			q=bdquery()
			listarecuperada =q.listarrendicion(desde,hasta,registro,especie,cultivar,categoria,camp)
			totalfilas=len(listarecuperada)
			self.tb_rendicion.setRowCount(totalfilas)
		
				
			fila =0
			acum=0
			for i in listarecuperada:
				self.tb_rendicion.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
				self.tb_rendicion.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
				self.tb_rendicion.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
				self.tb_rendicion.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
				self.tb_rendicion.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
				self.tb_rendicion.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
				self.tb_rendicion.setItem(fila,6,QtGui.QTableWidgetItem(str(i[6])))
				self.tb_rendicion.setItem(fila,7,QtGui.QTableWidgetItem(str(i[7])))
				self.tb_rendicion.setItem(fila,8,QtGui.QTableWidgetItem(str(i[8])))
				self.tb_rendicion.setItem(fila,9,QtGui.QTableWidgetItem(str(i[9])))
				self.tb_rendicion.setItem(fila,10,QtGui.QTableWidgetItem(str(i[10])))
				fila=fila+1
				acum=acum+int(i[1])
				
			self.signal_total_rendicion.setText(str(acum))
			
		else:
			
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			
			if registro =="":
				registro="%"
			
			
			q=bdquery()
			listarecuperada =q.listarrendicionsolopedidos(desde,hasta,registro)
			totalfilas=len(listarecuperada)
			self.tb_rendicion.setRowCount(totalfilas)
			
			
			fila =0
			acum=0
			for i in listarecuperada:
				self.tb_rendicion.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
				self.tb_rendicion.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
				self.tb_rendicion.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
				self.tb_rendicion.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
				self.tb_rendicion.setItem(fila,4,QtGui.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,5,QtGui.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,6,QtGui.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,7,QtGui.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,8,QtGui.QTableWidgetItem("-"))
				self.tb_rendicion.setItem(fila,9,QtGui.QTableWidgetItem(str(i[4])))
				self.tb_rendicion.setItem(fila,10,QtGui.QTableWidgetItem(str(i[5])))
				
				fila=fila+1
				acum=acum+int(i[1])
			
			
			
			self.signal_total_rendicion.setText(str(acum))
				
			
		
	def exportarrendicion(self):
		
		if self.rb_todo_rendicion.isChecked():
			
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			especie=str(self.cbx_especie_rendicion.currentText())
			cultivar=str(self.cbx_cultivar_rendicion.currentText())
			categoria=str(self.cbx_categoria_rendicion.currentText())
			camp= str(self.cbx_camp_rendicion.currentText())
		
		
		
		
			if registro =="":
				registro="%"
			
			if especie =="":
				especie="%"
				
			if cultivar =="":
				cultivar="%"
		
			if categoria =="":
				categoria="%"
			
			if camp =="":
				camp="%"
			
				
			q=bdquery()
			tablarecuperada =q.listarrendicion(desde,hasta,registro,especie,cultivar,categoria,camp)
			book = Workbook()
			sheet = book.active
		
			sheet['A1']="RENDICION MENSUAL"
			sheet['A2']="PERIODO"
			sheet['B2']=desde +" a "+hasta	
			sheet['D2']="CANTIDAD ASIGNADA"	
			sheet['E2']="=SUM(B3:B1000)"
				
			sheet['A3']="RANGO"
			sheet['B3']="CANTIDAD"
			sheet['C3']="RNCYFS"
			sheet['D3']="RAZON SOCIAL"
			sheet['E3']="KG"
			sheet['F3']="ESPECIE"
			sheet['G3']="CULTIVAR"
			sheet['H3']="CATEGORIA"
			sheet['I3']="CAMPANA"
			sheet['J3']="FECHA"
			sheet['K3']="PEDIDO N"
		
			for i in tablarecuperada:
				sheet.append(i)
			
					
						
			
			book.save('rendicion_subpedidos.xls')
			
			
			
		else:
			desde =str(self.fechadesde_rendicion.text())
			hasta = str(self.fechahasta_rendicion.text())
			registro=str(self.cbx_rncyfs_rendicion.currentText())
			
			if registro =="":
				registro="%"
			
			
			q=bdquery()
			tablarecuperada =q.listarrendicionsolopedidos(desde,hasta,registro)
			book = Workbook()
			sheet = book.active
		
			sheet['A1']="RENDICION MENSUAL"
			sheet['A2']="PERIODO"
			sheet['B2']=desde +" a "+hasta
			sheet['D2']="CANTIDAD ENTREGADA"	
			sheet['E2']="=SUM(B3:B500)"
			
				
			sheet['A3']="RANGO"
			sheet['B3']="CANTIDAD"
			sheet['C3']="RNCYFS"
			sheet['D3']="RAZON SOCIAL"
			sheet['E3']="FECHA"
			sheet['F3']="PEDIDO N"
		
			for i in tablarecuperada:
				sheet.append(i)
			
					
						
			
			book.save('rendicion_pedidos.xls')
		
					
			
		
	def setearlockers(self):
		cantidad=self.txt_definir_locker.text()
		q=bdquery()
			
		for i in range(1,int(cantidad)):
			estado ="Disponible"
			q.definircantidadlockers(i,estado)
			
	def lockerdisponibles(self):
		q=bdquery()
		lockers=q.recuperalockers()
		
		#la consulta devuelve una tupla, por lo tanto hay que convertirla a str para llenar el combobox, se usa el metodo "".join()
		
		k=0
		
		for i in lockers:
			self.cbx_num_locker.addItem("".join(map(str,lockers[k]))) #uso funcion map para pasar de tupla entero a string
			k=k+1
		
	def almacenar(self):
		
		if self.txt_pedido_deposito.text():
			
			locker = str(self.cbx_num_locker.currentText())
			pedido =int(self.txt_pedido_deposito.text())
			fecha= date.today()
			q=bdquery()
			q.modificalocker(locker,pedido,fecha)
			
			self.cbx_num_locker.clear()
			self.lockerdisponibles()
			self.listarlockers()
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("INGRESE PEDIDO A ALMACENAR")
			msgBox.exec_()
		
		
	def lockerselected(self):
		fila = self.tb_lockers.currentRow()
		locker=self.tb_lockers.item(fila, 0).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		pedido=self.tb_lockers.item(fila, 1).text()
		nombre=self.tb_lockers.item(fila, 5).text()
		
		
		self.signal_pedido_locker.setText(str(pedido))
		self.signal_locker_locker.setText(str(locker))
		self.signal_razon_locker.setText(str(nombre))
			
		
	def despachar(self):
		locker = int(self.signal_locker_locker.text())
		
		
		q=bdquery()
		q.liberalocker(locker)
		
		self.cbx_num_locker.clear()
		self.lockerdisponibles()
		self.listarlockers()
		
		
			
	def listarlockers(self):
		q=bdquery()
		listarecuperada=q.verlockers()
		totalfilas=len(listarecuperada)
		self.tb_lockers.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_lockers.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_lockers.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_lockers.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_lockers.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_lockers.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_lockers.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			
				
			fila=fila+1
			
	def filtrarlockers(self):
		q=bdquery()
		locker=str("%"+self.txt_buscar_locker.text()+"%")
		listarecuperada=q.verlockers_filtrado(locker)
		totalfilas=len(listarecuperada)
		self.tb_lockers.setRowCount(totalfilas)
			
			
		fila =0
		
		for i in listarecuperada:
			self.tb_lockers.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_lockers.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_lockers.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_lockers.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_lockers.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_lockers.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			
				
			fila=fila+1
		
		
	def nuevaimpresion(self):
		
		if self.txt_cantidad_rotulos.text():
			registro=str(self.txt_rncyfs_rotulos.text())
			cantidad=int(self.txt_cantidad_rotulos.text())
			especie=str(self.cbx_especie_rotulos.currentText())
			categoria=str(self.cbx_categoria_rotulos.currentText())
			tipo=str(self.cbx_tipo_rotulos.currentText())
			fecha=str(date.today())
			estado="PENDIENTE"
			
			q=bdquery()
			q.altarotulo(registro,especie,tipo,cantidad,estado,categoria,fecha)
			
			self.txt_rncyfs_rotulos.setText("")
			self.txt_cantidad_rotulos.setText("")
			self.cbx_especie_rotulos.setCurrentIndex(0)
			self.cbx_categoria_rotulos.setCurrentIndex(0)
			self.cbx_tipo_rotulos.setCurrentIndex(0)
			self.listarimpresiones()
		else:
			msgBox=QtGui.QMessageBox(self.centralwidget)
			msgBox.setIcon(3)
			msgBox.setWindowTitle("ERROR")
			msgBox.setText("INGRESE CANTIDAD")
			msgBox.exec_()
		
		
		
	def listarimpresiones(self):
		
		if self.rb_rotulos_pendientes.isChecked():
			estado="PENDIENTE"
		elif self.rb_rotulos_facturados.isChecked():
			estado="FACTURADO"
		elif self.rb_rotulos_todos.isChecked():
			estado="%"
		
		
		if self.cb_tipo.currentText() == "-":
			tipo="%"
		else:
			
			tipo=str(self.cb_tipo.currentText())
		
		
		q=bdquery()
		
		listarecuperada=q.traerotulos(estado,tipo)
		totalfilas=len(listarecuperada)
		self.tb_rotulos.setRowCount(totalfilas)
			
			
		fila =0
		acum=0
		for i in listarecuperada:
			self.tb_rotulos.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_rotulos.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_rotulos.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_rotulos.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
			self.tb_rotulos.setItem(fila,4,QtGui.QTableWidgetItem(str(i[4])))
			self.tb_rotulos.setItem(fila,5,QtGui.QTableWidgetItem(str(i[5])))
			self.tb_rotulos.setItem(fila,6,QtGui.QTableWidgetItem(str(i[6])))
			self.tb_rotulos.setItem(fila,7,QtGui.QTableWidgetItem(str(i[7])))
			
			acum=acum+int(i[3])
				
			fila=fila+1
			
			self.signal_total_rotulos.setText(str(acum))
		
			
	def cambiarestadorotulo(self):
		indice=int(self.txt_indice_rotulos.text())
		estado=str(self.txt_estado_rotulos.currentText())
		q=bdquery()
		q.definirestadorotulo(estado,indice)
		self.listarimpresiones()
		
		
		
	def impresionselected(self):
		fila = self.tb_rotulos.currentRow()
		indice=self.tb_rotulos.item(fila, 0).text() #SELECCIONO EL CONTENIDO DE LA FILA 5 DE LA COLUMNA SELECCIONADA
		razon=self.tb_rotulos.item(fila, 4).text()
		cantidad=self.tb_rotulos.item(fila, 3).text()
		
		
		
		
		
		self.signal_rotulos_razon.setText(str(razon))
		self.signal_rotulos_cantidad.setText(str(cantidad))
		self.txt_indice_rotulos.setText(str(indice))
		
			
		
	def agregarrango(self):
			inicio=int(self.txt_inicio_nuevorango.text())
			cantidad=int(self.txt_fin_nuevorango.text())
			if self.rb_seriea_add.isChecked():
				serie="A"
			elif self.rb_serieb_add.isChecked():
				serie="B"
			fin=inicio+cantidad-1
			q=bdquery()
			q.nuevorango(inicio,fin,cantidad,serie)
			
			self.txt_inicio_nuevorango.setText("")
			self.txt_fin_nuevorango.setText("")
			
			
	def verrangos(self):
		
		if self.rb_rangos_disponibles.isChecked():
			estado="DISPONIBLE"
		elif self.rb_rangos_enuso.isChecked():
			estado="EN USO"
		elif self.rb_rangos_terminados.isChecked():
			estado="TERMINADOS"
		elif self.rb_rangos_todos.isChecked():
			estado="%"
		
		if self.rb_definir_seriea.isChecked():
			serie="A"
		else:
			serie="B"
		
		q=bdquery()
		listarecuperada=q.traerangos(estado,serie)
		totalfilas=len(listarecuperada)
		self.tb_rangos.setRowCount(totalfilas)
			
			
		fila =0
		acum=0
		for i in listarecuperada:
			self.tb_rangos.setItem(fila,0,QtGui.QTableWidgetItem(str(i[0])))
			self.tb_rangos.setItem(fila,1,QtGui.QTableWidgetItem(str(i[1])))
			self.tb_rangos.setItem(fila,2,QtGui.QTableWidgetItem(str(i[2])))
			self.tb_rangos.setItem(fila,3,QtGui.QTableWidgetItem(str(i[3])))
							
			fila=fila+1
			acum=acum+int(i[1])
		
		self.signal_total_rangos.setText(str(acum))
		
		
		
	def setearrango(self):
		inicio=int(self.txt_inicioa_config.text())
		final=int(self.txt_final_config.text())
		if self.rb_definir_seriea.isChecked():
			indice=1
		else:
			indice=2
		q=bdquery()
		q.definirrango(inicio,final,indice)
		
		self.txt_inicioa_config.setText("")
		self.txt_final_config.setText("")
		
		

		
def salir():
	exit()
		




		
class Pedido:
	
		def __init__(self,cantidad,rncyfs,num):
			self.cantidad=cantidad
			self.rncyfs=rncyfs
			self.numpedido=num
		
		        
		def asignar(self,inicio):
		
			self.inicio=inicio
			self.fin=inicio+self.cantidad-1
			self.disponibleinicio=inicio
			self.disponiblefin=inicio+self.cantidad-1
		
		def showrango(self):
			print "Rango asignado:",self.inicio," - ",self.fin		
		
		
		
def listarsub():
	for i in subpedidos:
		print "pedido num: ",i.numpedido, "Cantidad Otorgada: ",i.cantidad,"Rango: ",i.inicio," - ",i.fin		
		
	
		
		
class Subpedido():
	def __init__(self,inicio,fin,cantidad,numpedido,variedad,especie,camp,dav,categoria,registro):
		self.numpedido= numpedido
		self.cantidad=cantidad
		self.inicio=inicio
		self.fin=fin
		self.variedad=variedad
		self.especie= especie
		self.categoria= categoria
		self.camp= camp
		self.dav=dav
		self.registro=registro


		

def consultastock():
	print "STOCK DISPONIBLE: ",(FINAL-INICIAL+1)


	
		
def actualizarangogeneral(cantidad):
	pass
	#Rango[0]=Rango[0]+cantidad
	INICIAL= INICIAL+cantidad
	q=bdquery()
					
	


if __name__ == '__main__':
	app = QtGui.QApplication(sys.argv)
	MyWindow = VentanaPrincipal(None)
	MyWindow.llenarcombo()
	MyWindow.lockerdisponibles()
	MyWindow.show()
	app.exec_()
	
	




		

   

    


